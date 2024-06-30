from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .models import Good, Category, TemplateFinalText
from io import BytesIO
import openpyxl

# Create your views here.
@login_required
def dashboard(request):
    goods = Good.objects.all()
    categories = Category.objects.all()
    template = TemplateFinalText.objects.get(pk=1).text #только один шаблон
    context = {
        'goods': goods,
        'categories': categories,
        'template' : template
    }
    return render(request, 'dashboard_index.html', context)

@require_POST
@login_required
def delete_good(request):
    good_id = request.POST.get('id')
    try:
        good = Good.objects.get(id=good_id)
        good.delete()
        return JsonResponse({'success': True})
    except Good.DoesNotExist:
        return JsonResponse({'success': False})
    
@require_GET   
def get_good_details(request, good_id):
    try:
        good = get_object_or_404(Good, id=good_id)
        data = {
            'good_name': good.good_name,
            'category_id': good.category_id,
            'catalog_cost': good.catalog_cost,
            'pv_value': good.pv_value
        }
        return JsonResponse(data)
    except Http404:
        return JsonResponse({'error': 'Good not found'}, status=404)

@require_POST
@login_required
def update_good(request):
    if request.method == 'POST':
        good_id = request.POST.get('good_id')
        good = get_object_or_404(Good, pk=good_id)  
        
        good.good_name = request.POST.get('editName')
        category_id = request.POST.get('editCategory')
        # ссылаемся на объект
        category = get_object_or_404(Category, pk=category_id)
        good.category = category
        good.catalog_cost = request.POST.get('editCost')
        good.pv_value = request.POST.get('editPV')
        
        # Сохраняем обновленный товар
        good.save()

        #Возвращаем дату для перерисовки фронта
        updated_good = Good.objects.get(pk=good_id)
        categories = Category.objects.all()
        goods = Good.objects.all()

        response_data = {
            'message': 'Товар успешно обновлен',
            'updated_good':{
                'good_name': updated_good.good_name,
                'category_id': updated_good.category.id,
                'catalog_cost': updated_good.catalog_cost,
                'pv_value': updated_good.pv_value,
            },
            'categories':list(categories.values('id','category')),
            'goods':list(goods.values('id','category_id','good_name','catalog_cost','pv_value'))
        }
        
        return JsonResponse(response_data, status=200)
    
    return JsonResponse({'error': 'Метод не разрешен'}, status=405)

def get_categories(request):
    if request.method == 'GET':
        categories = Category.objects.all()
        categories_list = list(categories.values('id', 'category'))
        return JsonResponse(categories_list, safe=False)
    return JsonResponse({'error': 'Метод не разрешен'}, status=405)

@require_POST
@login_required
def delete_category(request):
    category_id = request.POST.get('category_id')
    if category_id:
        category = get_object_or_404(Category, pk=category_id)
        push_category_name = category.category
        category.delete()

        categories = Category.objects.all()
        goods = Good.objects.all()

        response_data = {
        'message': 'Категория ' + push_category_name +  ' удалена',
        'categories':list(categories.values('id','category')),
        'goods':list(goods.values('id','category_id','good_name','catalog_cost','pv_value'))
        }
        return JsonResponse(response_data, status=200)
    return JsonResponse({'error': 'ID категории не предоставлен'}, status=400)

@require_POST
@login_required
def add_good(request):
    good_name = request.POST.get('good_name')
    category_id = request.POST.get('category_id')
    catalog_cost = request.POST.get('catalog_cost')
    pv_value = request.POST.get('pv_value')

    category = get_object_or_404(Category, pk=category_id)
    
    good = Good.objects.create(
        good_name=good_name,
        category=category,
        catalog_cost=catalog_cost,
        pv_value=pv_value
    )
    good.save()
    categories = Category.objects.all()
    goods = Good.objects.all()

    response_data = {
        'message': 'Товар \"' + good_name + '\" успешно добавлен',
        'categories': list(categories.values('id', 'category')),
        'goods': list(goods.values('id', 'category_id', 'good_name', 'catalog_cost', 'pv_value'))
    }

    return JsonResponse(response_data, status=200)

@csrf_exempt
@login_required
def add_category(request):
    if request.method == 'POST':
        category_name = request.POST.get('category_name')
        if category_name:
            category = Category.objects.create(category=category_name)
            categories = Category.objects.all().values('id', 'category')
            goods = Good.objects.all().values('id', 'good_name', 'category_id', 'catalog_cost', 'pv_value')
            return JsonResponse({
                'success': True,
                'message': 'Категория \"' + category.category + '\" успешно добавлена',
                'categories': list(categories),
                'goods': list(goods)
            })
        else:
            return JsonResponse({'success': False, 'message': 'Название категории не может быть пустым'}, status=400)

    return JsonResponse({'success': False, 'message': 'Неподдерживаемый метод'}, status=405)

@csrf_exempt
@login_required
def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            category_name, good_name, catalog_cost, pv_value = row
            category, created = Category.objects.get_or_create(category=category_name)
            
            # Проверяем, существует ли товар с таким именем и категорией
            good, good_created = Good.objects.get_or_create(
                good_name=good_name,
                category=category,
                defaults={
                    'catalog_cost': catalog_cost,
                    'pv_value': pv_value
                }
            )    
            # Если товар существует, обновляем его оставшиеся данные
            if not good_created:
                good.catalog_cost = catalog_cost
                good.pv_value = pv_value
                good.save()
        
        categories = Category.objects.all().values('id', 'category')
        goods = Good.objects.all().values('id', 'good_name', 'category_id', 'catalog_cost', 'pv_value')
        return JsonResponse({
            'success': True,
            'message': 'Файл успешно загружен и обработан',
            'categories': list(categories),
            'goods': list(goods)
        })
    return JsonResponse({'success': False, 'message': 'Неподдерживаемый метод'}, status=405)

@login_required
def export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    # Заполнение заголовков столбцов
    headers = ['Категория', 'Наименование', 'Цена по каталогу', 'Объем в PV']
    ws.append(headers)

    # Получение данных из базы данных и заполнение строк
    categories = Category.objects.all()
    for category in categories:
        goods = Good.objects.filter(category=category)
        for good in goods:
            row = [
                category.category,
                good.good_name,
                good.catalog_cost,
                good.pv_value
            ]
            ws.append(row)

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    # Создание HTTP ответа с типом содержимого, соответствующим Excel файлу
    response = HttpResponse(content=virtual_workbook, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=gw_export_goods.xlsx'
    return response

@csrf_exempt
@login_required
def update_template(request):
    if request.method == 'POST':
        template = TemplateFinalText.objects.first()
        if template:
            new_text = request.POST.get('text')
            if new_text:
                template.text = new_text
                template.save()
            return JsonResponse({
                'success': True,
                'message': 'Шаблон успешно обновлен'
            })
        else:
            return JsonResponse({'success': False, 'message': 'Ошибка!'}, status=400)

    return JsonResponse({'success': False, 'message': 'Неподдерживаемый метод'}, status=405)